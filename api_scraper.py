import requests
import pandas as pd
import time
import os

# 📂 Dynamically calculate the global path to 'outputs/'
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_PATH = os.path.join(BASE_DIR, "outputs", "multi_page_quotes.xlsx")

# 📥 Our global data basket
all_quotes = []

current_page = 1
has_next_page = True

print("🚀 Starting Multi-Page API Engine...")

while has_next_page:
    print(f"📡 Fetching Page {current_page}...")
    
    # Dynamically inject the page number directly into the URL string
    url = f"https://quotes.toscrape.com/api/quotes?page={current_page}"
    
    response = requests.get(url, timeout=20)
    
    # 🌟 Convert response text to Python Dictionary
    data = response.json()
    
    # Extract the array of quotes
    quotes_list = data["quotes"]
    
    # Loop through the quotes on this page and extract what we need
    for item in quotes_list:
        quote_text = item["text"]
        author_name = item["author"]["name"]
        
        all_quotes.append({
            "Quote": quote_text,
            "Author": author_name
        })
    
    # 🔁 Check if we should continue looping
    has_next_page = data["has_next"]
    
    if has_next_page:
        current_page += 1  # Increment page number
        time.sleep(1)     # Polite 1-second delay so we don't spam the server
    else:
        print("🏁 Hit the final page. Stopping loop.")

# 📊 Convert our list of dictionaries into a clean DataFrame structure
df = pd.DataFrame(all_quotes)
print(f"🎉 Success! Extracted {len(all_quotes)} rows across all pages.")

print("📊 Styling and formatting the Excel workbook...")

# 🌟 FIXED PATH: We pass OUTPUT_PATH directly to the ExcelWriter here!
with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Scraped Quotes", index=False)
    
    # Access the underlying workbook object to apply direct cosmetics
    workbook = writer.book
    worksheet = writer.sheets["Scraped Quotes"]
    
    # Import styling tools from openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Define our premium color palette styles
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Dark Charcoal
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")               # Crisp White
    data_font = Font(name="Segoe UI", size=10, color="333333")                            # Soft Off-Black
    
    # Define a clean, subtle light grey border rule
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Format the Header Row (Row 1)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Format all Data Rows
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.font = data_font
            cell.border = grid_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            
    # Set custom proportions so columns stretch elegantly
    worksheet.column_dimensions["A"].width = 75  # Massive room for Quotes
    worksheet.column_dimensions["B"].width = 25  # Neat breathing room for Author
    
    # Set row heights so things feel spacious
    worksheet.row_dimensions[1].height = 28  # Taller header row
    for r in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[r].height = 40  # Comfortable breathing room

print(f"🎉 Process Complete! Clean, styled file saved to: {OUTPUT_PATH}")